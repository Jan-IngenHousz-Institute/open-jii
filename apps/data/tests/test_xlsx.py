"""Tests for exports.xlsx — pure helpers used by data_export_task.py."""

import json

import pandas as pd
import pytest
from openpyxl import load_workbook

from exports.xlsx import (
    EXCEL_CELL_CHAR_LIMIT,
    EXCEL_SHEET_NAME_LIMIT,
    sanitize_sheet_name,
    truncate_long_strings,
    write_xlsx,
)


class TestSanitizeSheetName:
    def test_keeps_simple_name(self):
        assert sanitize_sheet_name("raw_data") == "raw_data"

    def test_falls_back_when_empty(self):
        assert sanitize_sheet_name("") == "data"
        assert sanitize_sheet_name(None) == "data"

    def test_truncates_to_31_chars(self):
        long = "a" * 50
        result = sanitize_sheet_name(long)
        assert len(result) == EXCEL_SHEET_NAME_LIMIT
        assert result == "a" * EXCEL_SHEET_NAME_LIMIT

    def test_replaces_disallowed_characters(self):
        # Excel forbids : \ / ? * [ ]
        assert sanitize_sheet_name("a:b/c\\d?e*f[g]h") == "a_b_c_d_e_f_g_h"

    def test_falls_back_when_only_invalid_chars_after_truncation(self):
        # Edge case: name reduces to empty after truncate+replace shouldn't happen
        # in practice since replacements keep length, but guard anyway.
        assert sanitize_sheet_name("\0") != ""


class TestTruncateLongStrings:
    def test_truncates_oversized_string_cell(self):
        long_value = "x" * (EXCEL_CELL_CHAR_LIMIT + 100)
        pdf = pd.DataFrame({"col": [long_value]})

        truncate_long_strings(pdf)

        assert len(pdf.iloc[0]["col"]) == EXCEL_CELL_CHAR_LIMIT

    def test_leaves_short_strings_alone(self):
        pdf = pd.DataFrame({"col": ["short"]})
        truncate_long_strings(pdf)
        assert pdf.iloc[0]["col"] == "short"

    def test_leaves_numeric_columns_alone(self):
        pdf = pd.DataFrame({"n": [1, 2, 3], "f": [1.5, 2.5, 3.5]})
        truncate_long_strings(pdf)
        assert pdf["n"].tolist() == [1, 2, 3]
        assert pdf["f"].tolist() == [1.5, 2.5, 3.5]

    def test_leaves_none_values_alone(self):
        pdf = pd.DataFrame({"col": ["ok", None, "also ok"]})
        truncate_long_strings(pdf)
        assert pdf.iloc[0]["col"] == "ok"
        assert pdf.iloc[1]["col"] is None
        assert pdf.iloc[2]["col"] == "also ok"

    def test_respects_custom_limit(self):
        pdf = pd.DataFrame({"col": ["abcdefgh"]})
        truncate_long_strings(pdf, limit=3)
        assert pdf.iloc[0]["col"] == "abc"


class TestWriteXlsx:
    def test_round_trips_simple_dataframe(self, tmp_path):
        pdf = pd.DataFrame(
            {
                "id": [1, 2, 3],
                "name": ["alpha", "beta", "gamma"],
                "ratio": [0.1, 0.2, 0.3],
            }
        )
        out = str(tmp_path / "out.xlsx")

        write_xlsx(pdf, out, "raw_data")

        wb = load_workbook(out)
        assert wb.sheetnames == ["raw_data"]
        ws = wb["raw_data"]
        rows = list(ws.iter_rows(values_only=True))
        assert rows[0] == ("id", "name", "ratio")
        assert rows[1] == (1, "alpha", 0.1)
        assert rows[3] == (3, "gamma", 0.3)

    def test_preserves_numeric_types(self, tmp_path):
        pdf = pd.DataFrame({"n": [42], "f": [3.14]})
        out = str(tmp_path / "types.xlsx")

        write_xlsx(pdf, out, "t")

        ws = load_workbook(out)["t"]
        # Header in row 1, data in row 2
        assert isinstance(ws.cell(row=2, column=1).value, int)
        assert isinstance(ws.cell(row=2, column=2).value, float)

    def test_preserves_timestamp_columns(self, tmp_path):
        pdf = pd.DataFrame({"ts": pd.to_datetime(["2026-01-15T12:30:00"])})
        out = str(tmp_path / "ts.xlsx")

        write_xlsx(pdf, out, "t")

        ws = load_workbook(out)["t"]
        from datetime import datetime
        assert isinstance(ws.cell(row=2, column=1).value, datetime)

    def test_serialized_json_stays_in_one_cell(self, tmp_path):
        # Mirrors what the task does with struct/array/map columns:
        # convert to JSON string before writing.
        json_str = json.dumps({"nested": {"a": 1, "b": [1, 2, 3]}})
        pdf = pd.DataFrame({"meta": [json_str]})
        out = str(tmp_path / "json.xlsx")

        write_xlsx(pdf, out, "t")

        ws = load_workbook(out)["t"]
        # JSON ends up in a single cell, recoverable by parsing.
        cell_value = ws.cell(row=2, column=1).value
        assert json.loads(cell_value) == {"nested": {"a": 1, "b": [1, 2, 3]}}

    def test_truncates_oversize_cells_so_workbook_is_valid(self, tmp_path):
        oversized = "y" * (EXCEL_CELL_CHAR_LIMIT + 500)
        pdf = pd.DataFrame({"big": [oversized]})
        out = str(tmp_path / "big.xlsx")

        write_xlsx(pdf, out, "t")

        ws = load_workbook(out)["t"]
        cell = ws.cell(row=2, column=1).value
        assert len(cell) == EXCEL_CELL_CHAR_LIMIT

    def test_sanitizes_sheet_name(self, tmp_path):
        out = str(tmp_path / "s.xlsx")

        write_xlsx(pd.DataFrame({"a": [1]}), out, "raw/data:2026")

        wb = load_workbook(out)
        assert wb.sheetnames == ["raw_data_2026"]

    def test_caps_sheet_name_length(self, tmp_path):
        out = str(tmp_path / "long.xlsx")
        long_name = "a" * 80

        write_xlsx(pd.DataFrame({"a": [1]}), out, long_name)

        wb = load_workbook(out)
        assert len(wb.sheetnames[0]) == EXCEL_SHEET_NAME_LIMIT

    def test_handles_empty_dataframe(self, tmp_path):
        # No rows, but the workbook should still be writable so the task
        # does not crash on empty exports (though task short-circuits earlier).
        out = str(tmp_path / "empty.xlsx")

        write_xlsx(pd.DataFrame({"a": [], "b": []}), out, "t")

        ws = load_workbook(out)["t"]
        rows = list(ws.iter_rows(values_only=True))
        assert rows == [("a", "b")]
