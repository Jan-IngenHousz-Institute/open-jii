"""Integration tests for exports.xlsx.

Exercises ``truncate_string_columns``, ``prepare_for_excel`` and
``write_xlsx`` against a local SparkSession. Skipped when pyspark is
not importable, or when a local Spark session can't be started (e.g.
no compatible JDK on the dev machine).
"""

from __future__ import annotations

import json

import pytest

pytest.importorskip("pyspark", reason="pyspark not installed locally")

from openpyxl import load_workbook  # noqa: E402
from pyspark.sql import SparkSession  # noqa: E402
from pyspark.sql.types import (  # noqa: E402
    ArrayType,
    IntegerType,
    StringType,
    StructField,
    StructType,
)

from exports.xlsx import (  # noqa: E402
    EXCEL_CELL_CHAR_LIMIT,
    EXCEL_SHEET_NAME_LIMIT,
    ExcelRowLimitError,
    prepare_for_excel,
    truncate_string_columns,
    write_xlsx,
)


@pytest.fixture(scope="module")
def spark():
    try:
        session = (
            SparkSession.builder.master("local[2]")
            .appName("exports-xlsx-tests")
            .config("spark.sql.shuffle.partitions", "2")
            .getOrCreate()
        )
    except Exception as exc:
        pytest.skip(f"Could not start a local Spark session: {exc}")
    yield session
    session.stop()


class TestTruncateStringColumns:
    def test_caps_oversize_string_cells(self, spark):
        oversized = "x" * (EXCEL_CELL_CHAR_LIMIT + 500)
        sdf = spark.createDataFrame([(oversized,)], ["s"])

        result = truncate_string_columns(sdf).collect()

        assert len(result[0]["s"]) == EXCEL_CELL_CHAR_LIMIT

    def test_leaves_short_strings_alone(self, spark):
        sdf = spark.createDataFrame([("hello",)], ["s"])
        assert truncate_string_columns(sdf).collect()[0]["s"] == "hello"

    def test_leaves_non_string_columns_alone(self, spark):
        sdf = spark.createDataFrame([(1, 2.5, "ok")], ["i", "f", "s"])
        row = truncate_string_columns(sdf).collect()[0]
        assert row["i"] == 1
        assert row["f"] == 2.5
        assert row["s"] == "ok"

    def test_respects_custom_limit(self, spark):
        sdf = spark.createDataFrame([("abcdefgh",)], ["s"])
        assert truncate_string_columns(sdf, limit=3).collect()[0]["s"] == "abc"


class TestPrepareForExcel:
    def test_struct_column_becomes_json_string(self, spark):
        schema = StructType(
            [
                StructField(
                    "meta",
                    StructType(
                        [
                            StructField("a", IntegerType()),
                            StructField("b", StringType()),
                        ]
                    ),
                ),
            ]
        )
        sdf = spark.createDataFrame([({"a": 1, "b": "two"},)], schema)

        cell = prepare_for_excel(sdf).collect()[0]["meta"]

        assert isinstance(cell, str)
        assert json.loads(cell) == {"a": 1, "b": "two"}

    def test_array_column_becomes_json_string(self, spark):
        schema = StructType([StructField("xs", ArrayType(IntegerType()))])
        sdf = spark.createDataFrame([([1, 2, 3],)], schema)

        cell = prepare_for_excel(sdf).collect()[0]["xs"]
        assert json.loads(cell) == [1, 2, 3]

    def test_string_columns_are_truncated(self, spark):
        oversized = "y" * (EXCEL_CELL_CHAR_LIMIT + 100)
        sdf = spark.createDataFrame([(oversized,)], ["s"])

        cell = prepare_for_excel(sdf).collect()[0]["s"]
        assert len(cell) == EXCEL_CELL_CHAR_LIMIT


class TestWriteXlsx:
    def test_round_trips_simple_dataframe(self, spark, tmp_path):
        sdf = spark.createDataFrame(
            [(1, "alpha", 0.1), (2, "beta", 0.2), (3, "gamma", 0.3)],
            ["id", "name", "ratio"],
        )
        out = str(tmp_path / "out.xlsx")

        report = write_xlsx(sdf, out, "raw_data")

        assert report == {"row_count": 3, "sheet_name": "raw_data", "output_path": out}

        wb = load_workbook(out)
        assert wb.sheetnames == ["raw_data"]
        rows = list(wb["raw_data"].iter_rows(values_only=True))
        assert rows[0] == ("id", "name", "ratio")
        assert rows[1][1] == "alpha"
        assert rows[3][0] == 3

    def test_serialized_struct_stays_in_one_cell(self, spark, tmp_path):
        schema = StructType(
            [
                StructField(
                    "nested",
                    StructType([StructField("k", StringType())]),
                )
            ]
        )
        sdf = spark.createDataFrame([({"k": "v"},)], schema)
        out = str(tmp_path / "nested.xlsx")

        write_xlsx(sdf, out, "t")

        ws = load_workbook(out)["t"]
        assert json.loads(ws.cell(row=2, column=1).value) == {"k": "v"}

    def test_truncates_oversize_cells_via_spark_preprocessing(self, spark, tmp_path):
        oversized = "y" * (EXCEL_CELL_CHAR_LIMIT + 500)
        sdf = spark.createDataFrame([(oversized,)], ["big"])
        out = str(tmp_path / "big.xlsx")

        write_xlsx(sdf, out, "t")

        ws = load_workbook(out)["t"]
        assert len(ws.cell(row=2, column=1).value) == EXCEL_CELL_CHAR_LIMIT

    def test_sanitizes_sheet_name(self, spark, tmp_path):
        sdf = spark.createDataFrame([(1,)], ["a"])
        out = str(tmp_path / "s.xlsx")

        report = write_xlsx(sdf, out, "raw/data:2026")

        assert report["sheet_name"] == "raw_data_2026"
        assert load_workbook(out).sheetnames == ["raw_data_2026"]

    def test_caps_sheet_name_length(self, spark, tmp_path):
        sdf = spark.createDataFrame([(1,)], ["a"])
        out = str(tmp_path / "long.xlsx")

        report = write_xlsx(sdf, out, "a" * 80)

        assert len(report["sheet_name"]) == EXCEL_SHEET_NAME_LIMIT

    def test_raises_when_row_count_exceeds_max(self, spark, tmp_path):
        sdf = spark.createDataFrame([(i,) for i in range(5)], ["i"])
        out = str(tmp_path / "too_big.xlsx")

        with pytest.raises(ExcelRowLimitError) as exc_info:
            write_xlsx(sdf, out, "t", max_rows=3)

        assert exc_info.value.row_count == 5
        assert exc_info.value.limit == 3
        assert not (tmp_path / "too_big.xlsx").exists()

    def test_handles_empty_dataframe(self, spark, tmp_path):
        sdf = spark.createDataFrame([], "a INT, b STRING")
        out = str(tmp_path / "empty.xlsx")

        report = write_xlsx(sdf, out, "t")

        assert report["row_count"] == 0
        ws = load_workbook(out)["t"]
        rows = list(ws.iter_rows(values_only=True))
        assert rows == [("a", "b")]
